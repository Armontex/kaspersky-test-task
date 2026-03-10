import pytest
from openpyxl import load_workbook
from src.services.analyzer.infra.writers.xlsx import ExcelWriter
from src.services.analyzer.domain.models import Result


class TestExcelWriter:

    @pytest.fixture
    def writer(self):
        return ExcelWriter()

    def test_create_report_content(self, tmp_path, writer):
        output_path = str(tmp_path / "report.xlsx")
        results = [
            Result(lemma="кот", global_count=5, count_from_lines="1,0,4"),
            Result(lemma="пёс", global_count=2, count_from_lines="0,2,0"),
        ]

        writer.create_report(results, output_path)

        wb = load_workbook(output_path)
        sheet = wb["Анализ"]

        assert sheet.cell(row=1, column=1).value == "Словоформа"
        assert sheet.cell(row=1, column=2).value == "Общее количество"
        assert sheet.cell(row=1, column=3).value == "Статистика по строкам"

        assert sheet.cell(row=2, column=1).value == "кот"
        assert sheet.cell(row=2, column=2).value == 5
        assert sheet.cell(row=2, column=3).value == "1,0,4"

        assert sheet.cell(row=3, column=1).value == "пёс"
        assert sheet.cell(row=3, column=2).value == 2
        assert sheet.cell(row=3, column=3).value == "0,2,0"
